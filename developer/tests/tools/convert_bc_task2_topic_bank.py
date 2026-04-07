#!/usr/bin/env python3
"""Convert the BC Task 2 workbook into stable JSON artifacts.

Outputs:
1. A canonical catalog that preserves raw workbook metadata.
2. `electron/resources/default-topics.json`, which the runtime already knows how to seed.
"""

from __future__ import annotations

import json
from collections import Counter, OrderedDict
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[3]
SOURCE_XLSX = ROOT / "2024.12.1-2025.1.31 BC机考大作文机经整理by橙.xlsx"
CANONICAL_OUTPUT = ROOT / "assets/generated/writing-topics/bc-task2-2024-12_2025-01.catalog.json"
DEFAULT_TOPICS_OUTPUT = ROOT / "electron/resources/default-topics.json"

PROMPT_TYPE_MAP = {
    "agree": "agree_disagree",
    "Discuss Both Sides": "discuss_both_views",
    "Double Question": "double_question",
    "positive": "positive_negative",
    "advantage": "advantages_disadvantages",
}

# The runtime topic schema only supports 8 Task 2 categories today.
# Keep a current-compatible derived category while preserving raw themes in the catalog.
THEME_CATEGORY_MAP = {
    "教育类": "education",
    "科技类": "technology",
    "生活类": "society",
    "环境类": "environment",
    "媒体类": "culture",
    "语言类": "culture",
    "政府类": "government",
    "经济类": "economy",
}

DIFFICULTY_MAP = {
    "★": 1,
    "★★": 2,
    "★★★": 3,
    "★★★★": 4,
    "★★★★★": 5,
}

HEADER = (
    "Essay Question",
    "Question Type",
    "Theme",
    "Secondary Theme",
    "Difficulty",
    "DATE",
    "Score",
    "Task Response",
    "Coherence and Cohesion",
    "Lexical Resource",
    "Grammatical Range and Accuracy",
)


@dataclass
class SourceOccurrence:
    sheet_row: int
    raw_question_type: str | None
    raw_theme: str | None
    raw_secondary_theme: str | None
    raw_difficulty: str | None
    raw_date: Any
    parsed_date_iso: str | None
    date_confidence: str | None

    def to_dict(self) -> dict[str, Any]:
        return {
            "sheet_row": self.sheet_row,
            "raw_question_type": self.raw_question_type,
            "raw_theme": self.raw_theme,
            "raw_secondary_theme": self.raw_secondary_theme,
            "raw_difficulty": self.raw_difficulty,
            "raw_date": self.raw_date,
            "parsed_date_iso": self.parsed_date_iso,
            "date_confidence": self.date_confidence,
        }


def normalize_prompt(text: str) -> str:
    return " ".join(text.split())


def to_tiptap_text_doc(text: str) -> str:
    return json.dumps(
        {
            "type": "doc",
            "content": [
                {
                    "type": "paragraph",
                    "content": [{"type": "text", "text": text}],
                }
            ],
        },
        ensure_ascii=False,
        separators=(",", ":"),
    )


def parse_sparse_excel_date(value: Any) -> tuple[str | None, str | None]:
    if value in (None, ""):
        return None, None

    if isinstance(value, (int, float)):
        # The column is stored as a raw serial under "General" format in the workbook.
        # Preserve it, but mark the parsed date as low confidence.
        base = datetime(1899, 12, 30)
        parsed = (base + timedelta(days=float(value))).date().isoformat()
        return parsed, "low"

    if isinstance(value, datetime):
        return value.date().isoformat(), "medium"

    return str(value), "low"


def map_theme(raw_theme: str | None) -> str | None:
    if raw_theme in (None, ""):
        return None
    try:
        return THEME_CATEGORY_MAP[raw_theme]
    except KeyError as exc:
        raise ValueError(f"Unmapped raw theme: {raw_theme}") from exc


def map_prompt_type(raw_prompt_type: str | None) -> str | None:
    if raw_prompt_type in (None, ""):
        return None
    try:
        return PROMPT_TYPE_MAP[raw_prompt_type]
    except KeyError as exc:
        raise ValueError(f"Unmapped prompt type: {raw_prompt_type}") from exc


def map_difficulty(raw_difficulty: str | None) -> int | None:
    if raw_difficulty in (None, ""):
        return None
    try:
        return DIFFICULTY_MAP[raw_difficulty]
    except KeyError as exc:
        raise ValueError(f"Unmapped difficulty value: {raw_difficulty}") from exc


def build_catalog() -> tuple[dict[str, Any], list[dict[str, Any]]]:
    workbook = load_workbook(SOURCE_XLSX, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))

    if tuple(rows[0]) != HEADER:
        raise ValueError(f"Unexpected workbook header: {rows[0]!r}")

    deduped: OrderedDict[str, dict[str, Any]] = OrderedDict()
    raw_prompt_count = Counter()
    raw_theme_count = Counter()
    raw_question_type_count = Counter()

    for sheet_row, row in enumerate(rows[1:], start=2):
        prompt = (row[0] or "").strip()
        if not prompt:
            continue

        raw_prompt_count[prompt] += 1
        raw_question_type_count[row[1]] += 1
        raw_theme_count[row[2]] += 1

        normalized_prompt = normalize_prompt(prompt)
        prompt_type = map_prompt_type(row[1])
        category = map_theme(row[2])
        secondary_category = map_theme(row[3])
        difficulty = map_difficulty(row[4])
        parsed_date_iso, date_confidence = parse_sparse_excel_date(row[5])

        occurrence = SourceOccurrence(
            sheet_row=sheet_row,
            raw_question_type=row[1],
            raw_theme=row[2],
            raw_secondary_theme=row[3],
            raw_difficulty=row[4],
            raw_date=row[5],
            parsed_date_iso=parsed_date_iso,
            date_confidence=date_confidence,
        )

        record = deduped.get(normalized_prompt)
        if record is None:
            record = {
                "source_id": f"bc-task2-{len(deduped) + 1:04d}",
                "type": "task2",
                "prompt": prompt,
                "prompt_normalized": normalized_prompt,
                "prompt_type": prompt_type,
                "prompt_type_raw": row[1],
                "category": category,
                "theme_raw": row[2],
                "secondary_category": secondary_category,
                "secondary_theme_raw": row[3],
                "difficulty": difficulty,
                "difficulty_raw": row[4],
                "tags": sorted(
                    {
                        value
                        for value in (row[1], row[2], row[3], row[4])
                        if value not in (None, "")
                    }
                ),
                "source_occurrences": [],
            }
            deduped[normalized_prompt] = record
        else:
            for field_name, new_value in (
                ("prompt_type", prompt_type),
                ("category", category),
                ("secondary_category", secondary_category),
                ("difficulty", difficulty),
            ):
                current_value = record[field_name]
                if current_value != new_value:
                    raise ValueError(
                        f"Conflicting duplicate rows for prompt '{prompt[:80]}...' in field "
                        f"{field_name}: {current_value!r} != {new_value!r}"
                    )

        record["source_occurrences"].append(occurrence.to_dict())

    topics = list(deduped.values())
    duplicates = {prompt: count for prompt, count in raw_prompt_count.items() if count > 1}
    direct_import = [
        {
            "type": topic["type"],
            "category": topic["category"],
            "difficulty": topic["difficulty"],
            "title_json": to_tiptap_text_doc(topic["prompt"]),
            "is_official": 1,
        }
        for topic in topics
    ]

    catalog = {
        "version": 1,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "source": {
            "workbook": SOURCE_XLSX.name,
            "sheet": sheet.title,
            "raw_row_count": len(rows) - 1,
            "unique_topic_count": len(topics),
            "duplicate_prompt_count": len(duplicates),
            "dedupe_strategy": "normalized prompt text",
            "prompt_type_mapping": PROMPT_TYPE_MAP,
            "theme_category_mapping": THEME_CATEGORY_MAP,
            "difficulty_mapping": DIFFICULTY_MAP,
            "raw_theme_distribution": dict(sorted(raw_theme_count.items())),
            "raw_prompt_type_distribution": dict(sorted(raw_question_type_count.items())),
            "notes": [
                "This workbook behaves like a Task 2 topic source table, not an essay-history table.",
                "Score-related columns are empty and are intentionally not transformed into essay records.",
                "The DATE column is sparse and stored as raw serial/general values, so parsed ISO dates are low-confidence hints only.",
                "Some raw themes do not exist in the current runtime enum; the catalog preserves raw themes while the import file uses a derived compatible category.",
            ],
        },
        "duplicates": [
            {"prompt": prompt, "count": count}
            for prompt, count in sorted(duplicates.items(), key=lambda item: item[0])
        ],
        "topics": topics,
    }
    return catalog, direct_import


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)
        fh.write("\n")


def main() -> None:
    catalog, direct_import = build_catalog()
    write_json(CANONICAL_OUTPUT, catalog)
    write_json(DEFAULT_TOPICS_OUTPUT, direct_import)

    print(f"catalog: {CANONICAL_OUTPUT}")
    print(f"default_topics: {DEFAULT_TOPICS_OUTPUT}")
    print(f"unique topics: {catalog['source']['unique_topic_count']}")
    print(f"duplicate prompts removed: {catalog['source']['duplicate_prompt_count']}")


if __name__ == "__main__":
    main()
